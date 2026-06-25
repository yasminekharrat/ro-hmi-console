# modules/reports/report_builder.py
# ---------------------------------------------------------------------------
# Builds an .xlsx bilan (daily or weekly) from the telemetry history buffer.
# Returns the file path of the generated file.
#
# Sheet layout
# ─────────────
# Sheet 1 – "Récapitulatif"  : one stat row per tag (mean / min / max / σ)
# Sheet 2-N – one sheet per group defined in REPORT_TAGS
#              columns: Heure | <tag A> | <tag B> | …  (hourly averages)
#
# Dependencies: openpyxl (pip install openpyxl --break-system-packages)
# ---------------------------------------------------------------------------

import os
import time
import datetime
from collections import defaultdict

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

from config.report_config import REPORT_TAGS
from main.services import telemetry_history as hist


# ── Palette (Siemens-ish industrial blue / grey) ─────────────────────────────
CLR_HEADER_BG  = "1C3A5A"   # dark navy
CLR_HEADER_FG  = "FFFFFF"
CLR_GROUP_BG   = "2980B9"   # mid blue
CLR_GROUP_FG   = "FFFFFF"
CLR_ALT_ROW    = "EAF2FB"   # very light blue
CLR_WARN_BG    = "F9EBEA"   # light red for fault/alarm tags
CLR_BORDER     = "AACBE8"

THIN = Side(style="thin", color=CLR_BORDER)
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _cell_style(ws, row, col, value=None, bold=False, bg=None, fg="000000",
                align="left", num_fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Arial", bold=bold, color=fg, size=10)
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border    = BORDER
    if bg:
        c.fill = PatternFill("solid", start_color=bg)
    if num_fmt:
        c.number_format = num_fmt
    return c


def _header_row(ws, row, labels: list[str]):
    for col, label in enumerate(labels, 1):
        _cell_style(ws, row, col, label, bold=True,
                    bg=CLR_HEADER_BG, fg=CLR_HEADER_FG, align="center")


def _auto_width(ws, min_w=10, max_w=40):
    for col_cells in ws.columns:
        length = max((len(str(c.value or "")) for c in col_cells), default=min_w)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = \
            min(max(length + 2, min_w), max_w)


# ── Core builder ─────────────────────────────────────────────────────────────

def build_report(period: str, since: float, until: float,
                 output_dir: str = "/tmp") -> str:
    """
    Build an Excel report for *period* ("daily" or "weekly").
    *since* / *until* are Unix timestamps.
    Returns the absolute path of the .xlsx file.
    """
    date_str = datetime.datetime.fromtimestamp(since).strftime("%Y-%m-%d")
    if period == "weekly":
        end_str  = datetime.datetime.fromtimestamp(until).strftime("%Y-%m-%d")
        fname    = f"Bilan_Hebdomadaire_{date_str}_{end_str}.xlsx"
        title    = f"Bilan Hebdomadaire — {date_str} au {end_str}"
    else:
        fname    = f"Bilan_Journalier_{date_str}.xlsx"
        title    = f"Bilan Journalier — {date_str}"

    path = os.path.join(output_dir, fname)
    wb   = openpyxl.Workbook()

    # ── Sheet 1: Récapitulatif ────────────────────────────────────────────
    ws_recap = wb.active
    ws_recap.title = "Récapitulatif"

    # Title
    ws_recap.merge_cells("A1:H1")
    c = ws_recap["A1"]
    c.value     = title
    c.font      = Font(name="Arial", bold=True, size=14, color=CLR_HEADER_FG)
    c.fill      = PatternFill("solid", start_color=CLR_HEADER_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws_recap.row_dimensions[1].height = 28

    # Meta row
    generated_at = datetime.datetime.now().strftime("%d/%m/%Y %H:%M")
    ws_recap.merge_cells("A2:H2")
    c = ws_recap["A2"]
    c.value     = f"Généré le {generated_at}  •  Thermeco Industrie SCADA"
    c.font      = Font(name="Arial", italic=True, size=9, color="555555")
    c.alignment = Alignment(horizontal="right", vertical="center")

    # Column headers
    headers = ["Groupe", "Variable", "Unité", "Moyenne", "Min", "Max",
               "Écart-type σ", "Nb mesures"]
    _header_row(ws_recap, 3, headers)

    # Data rows — one per tag
    for i, tag_cfg in enumerate(REPORT_TAGS):
        row    = i + 4
        s      = hist.stats_for_tag(tag_cfg["tag_id"], since=since, until=until)
        is_alt = (i % 2 == 0)
        bg     = CLR_ALT_ROW if is_alt else "FFFFFF"
        # Highlight alarm/fault tags
        if "fault" in tag_cfg["tag_id"] or "alarme" in tag_cfg["tag_id"].lower():
            bg = CLR_WARN_BG

        unit_label = f"({tag_cfg['unit']})" if tag_cfg["unit"] else "—"
        prec       = tag_cfg.get("precision", 2)
        num_fmt    = f"0.{'0'*prec}" if prec > 0 else "0"

        _cell_style(ws_recap, row, 1, tag_cfg["group"], bg=bg, align="center")
        _cell_style(ws_recap, row, 2, tag_cfg["label"], bg=bg)
        _cell_style(ws_recap, row, 3, unit_label, bg=bg, align="center")
        _cell_style(ws_recap, row, 4, s["mean"],  bg=bg, align="right", num_fmt=num_fmt)
        _cell_style(ws_recap, row, 5, s["min"],   bg=bg, align="right", num_fmt=num_fmt)
        _cell_style(ws_recap, row, 6, s["max"],   bg=bg, align="right", num_fmt=num_fmt)
        _cell_style(ws_recap, row, 7, s["std"],   bg=bg, align="right", num_fmt=num_fmt)
        _cell_style(ws_recap, row, 8, s["count"], bg=bg, align="right", num_fmt="#,##0")

    _auto_width(ws_recap)
    ws_recap.freeze_panes = "A4"

    # ── Sheets 2-N: hourly averages per group ────────────────────────────
    groups: dict[str, list[dict]] = defaultdict(list)
    for t in REPORT_TAGS:
        groups[t["group"]].append(t)

    for group_name, tags in groups.items():
        ws = wb.create_sheet(title=group_name[:31])  # Excel tab name ≤31 chars

        # Title
        ncols = len(tags) + 1
        ws.merge_cells(start_row=1, start_column=1,
                       end_row=1, end_column=ncols)
        c = ws.cell(row=1, column=1, value=f"{title}  –  {group_name}")
        c.font      = Font(name="Arial", bold=True, size=12,
                           color=CLR_HEADER_FG)
        c.fill      = PatternFill("solid", start_color=CLR_GROUP_BG)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 22

        # Column headers
        col_headers = ["Heure"] + [
            f"{t['label']}" + (f" ({t['unit']})" if t['unit'] else "")
            for t in tags
        ]
        _header_row(ws, 2, col_headers)

        # Build hourly buckets between since → until
        bucket_hours = int((until - since) / 3600) + 1
        for h in range(bucket_hours):
            h_since = since + h * 3600
            h_until = h_since + 3600
            label   = datetime.datetime.fromtimestamp(h_since).strftime("%d/%m %H:00")
            row     = h + 3
            bg      = CLR_ALT_ROW if h % 2 == 0 else "FFFFFF"

            _cell_style(ws, row, 1, label, bg=bg, align="center")
            for col, tag_cfg in enumerate(tags, 2):
                s      = hist.stats_for_tag(tag_cfg["tag_id"],
                                            since=h_since, until=h_until)
                prec   = tag_cfg.get("precision", 2)
                num_fmt = f"0.{'0'*prec}" if prec > 0 else "0"
                _cell_style(ws, row, col, s["mean"], bg=bg,
                            align="right", num_fmt=num_fmt)

        _auto_width(ws)
        ws.freeze_panes = "A3"

    wb.save(path)
    return path